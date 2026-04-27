"""
Import ore reali per commessa da Excel (foglio Elaborato) e matching con nomi file offerte.
Le colonne Excel (IMBA, NEST, PIEG, PROD, PROG, SALD) allineano alle fasi del modello dove possibile;
NEST in Excel include taglio/nesting (nel modello: OreNest; OreTaglio resta 0 se non scomposto).
"""
from __future__ import annotations

import os
import re
from datetime import datetime, timezone
from typing import Any

import psycopg2
import psycopg2.extras


def _get_db_conn():
    url = os.environ.get('DATABASE_URL', '')
    conn = psycopg2.connect(url, connection_factory=psycopg2.extras.RealDictConnection)
    return conn

import openpyxl
import pandas as pd

# Colonne attese prima riga foglio "Elaborato"
HEADER_COMM = "Nr. Commessa"
COL_IMBA = "IMBA"
COL_NEST = "NEST"
COL_PIEG = "PIEG"
COL_PROD = "PROD"
COL_PROG = "PROG"
COL_SALD = "SALD"
COL_TOT = "Totale Ore"

_COMM_RE = re.compile(
    r"^\s*(?P<nr>\d{2}/\d{3})\s+(?:\"(?P<q>[^\"]+)\"|(?P<nq>.+))$",
    re.UNICODE,
)

_PREV_FN = re.compile(
    r"^(\d{4})_(.+?)_preventivo_(\d+)",
    re.IGNORECASE,
)


def default_xlsx_path() -> str:
    root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    candidates = [
        os.path.join(root, "dati", "ORE PER REPARTO PER COMMESSA commesse 25.xlsm"),
        os.path.join(root, "dati", "ORE_PER_REPARTO_commesse_25_Elaborato.csv"),
        os.path.join(root, "ORE PER REPARTO PER COMMESSA commesse 25.xlsm"),
    ]
    for p in candidates:
        if os.path.isfile(p):
            return p
    return candidates[0]


def normalize_cliente(s: str) -> str:
    if not s:
        return ""
    t = s.upper().strip()
    for suf in (
        " S.R.L.",
        " S.P.A.",
        " SRL",
        " SPA",
        " SAS",
        " S.A.S.",
        " SNC",
    ):
        if t.endswith(suf):
            t = t[: -len(suf)].strip()
    t = re.sub(r"[^A-Z0-9\s]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def parse_commessa_cell(raw: Any) -> tuple[str | None, str | None]:
    """Estrae nr commessa (es. 25/001) e ragione sociale dalla cella Excel."""
    if raw is None:
        return None, None
    s = str(raw).strip()
    if not s:
        return None, None
    m = _COMM_RE.match(s)
    if m:
        nr = m.group("nr")
        cliente = (m.group("q") or m.group("nq") or "").strip()
        return nr, cliente
    return None, None


def parse_preventivo_filename(filename: str) -> dict[str, Any] | None:
    """
    Da '2025_ARE_preventivo_12826.pdf' o '..._12964 conferma.pdf' estrae anno, cliente slug, numero.
    """
    base = os.path.basename(filename)
    stem = os.path.splitext(base)[0]
    stem = re.sub(r"\s+conferma\s*$", "", stem, flags=re.IGNORECASE)
    m = _PREV_FN.match(stem)
    if not m:
        return None
    anno, slug, num = m.group(1), m.group(2).strip(), m.group(3)
    cliente = slug.replace("_", " ").strip()
    return {
        "anno": anno,
        "cliente_slug": cliente,
        "cliente_norm": normalize_cliente(cliente),
        "numero_preventivo": num,
        "filename": base,
    }


def _canonical_header(cell_val: Any) -> str | None:
    """Mappa intestazioni tipo 'IMBA  \" IMBALLAGGIO\"' -> IMBA."""
    if cell_val is None:
        return None
    s = str(cell_val).strip().upper()
    if "COMMESSA" in s or s.startswith("NR."):
        return HEADER_COMM
    if s.startswith("IMBA"):
        return COL_IMBA
    if s.startswith("NEST"):
        return COL_NEST
    if s.startswith("PIEG"):
        return COL_PIEG
    if s.startswith("PROD"):
        return COL_PROD
    if s.startswith("PROG"):
        return COL_PROG
    if s.startswith("SALD"):
        return COL_SALD
    if "TOTALE" in s and "ORE" in s:
        return COL_TOT
    return None


def _header_map(ws) -> dict[str, int]:
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    mp: dict[str, int] = {}
    for i, v in enumerate(row1):
        key = _canonical_header(v)
        if key and key not in mp:
            mp[key] = i
    return mp


def load_rows_from_csv(path: str) -> list[dict[str, Any]]:
    """Legge il foglio Elaborato esportato in CSV (stessa struttura dell'Excel)."""
    df = pd.read_csv(path, encoding="utf-8-sig")
    col_map: dict[str, str] = {}
    for col in df.columns:
        key = _canonical_header(col)
        if key and key not in col_map.values():
            col_map[col] = key
    df = df.rename(columns=col_map)

    for k in (HEADER_COMM, COL_TOT):
        if k not in df.columns:
            raise ValueError(f"Colonna mancante '{k}'. Trovate: {list(df.columns)}")

    out: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        nr, cliente = parse_commessa_cell(row[HEADER_COMM])
        if not nr:
            continue

        def cell(name: str, _row=row) -> float | None:
            if name not in df.columns:
                return None
            v = _row[name]
            try:
                if pd.isna(v):
                    return None
            except (TypeError, ValueError):
                pass
            try:
                return float(v)
            except (TypeError, ValueError):
                return None

        out.append(
            {
                "nr_commessa": nr,
                "cliente": cliente or "",
                "cliente_norm": normalize_cliente(cliente or ""),
                "ore_imba": cell(COL_IMBA),
                "ore_nest": cell(COL_NEST),
                "ore_pieg": cell(COL_PIEG),
                "ore_prod": cell(COL_PROD),
                "ore_prog": cell(COL_PROG),
                "ore_sald": cell(COL_SALD),
                "ore_totale": cell(COL_TOT),
            }
        )
    return out


def load_rows_from_xlsx(path: str) -> list[dict[str, Any]]:
    if not os.path.isfile(path):
        raise FileNotFoundError(path)
    if path.lower().endswith(".csv"):
        return load_rows_from_csv(path)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Elaborato" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Foglio 'Elaborato' non trovato. Fogli: {wb.sheetnames}")

    ws = wb["Elaborato"]
    h = _header_map(ws)
    required = [HEADER_COMM, COL_TOT]
    for k in required:
        if k not in h:
            wb.close()
            raise ValueError(f"Colonna mancante '{k}'. Trovate: {list(h.keys())}")

    optional_cols = [COL_IMBA, COL_NEST, COL_PIEG, COL_PROD, COL_PROG, COL_SALD]
    out: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_comm = row[h[HEADER_COMM]]
        nr, cliente = parse_commessa_cell(raw_comm)
        if not nr:
            continue

        def cell(name: str) -> float | None:
            if name not in h:
                return None
            v = row[h[name]]
            if v is None or v == "":
                return None
            try:
                return float(v)
            except (TypeError, ValueError):
                return None

        tot = cell(COL_TOT)
        out.append(
            {
                "nr_commessa": nr,
                "cliente": cliente or "",
                "cliente_norm": normalize_cliente(cliente or ""),
                "ore_imba": cell(COL_IMBA),
                "ore_nest": cell(COL_NEST),
                "ore_pieg": cell(COL_PIEG),
                "ore_prod": cell(COL_PROD),
                "ore_prog": cell(COL_PROG),
                "ore_sald": cell(COL_SALD),
                "ore_totale": tot,
            }
        )
    wb.close()
    return out


def import_to_sqlite(
    db_path: str,
    rows: list[dict[str, Any]],
    source_file: str,
) -> dict[str, Any]:
    conn = _get_db_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM commesse_ore")
    now = datetime.now(timezone.utc).isoformat()
    for r in rows:
        cur.execute(
            """
            INSERT INTO commesse_ore (
                nr_commessa, cliente, cliente_norm,
                ore_imba, ore_nest, ore_pieg, ore_prod, ore_prog, ore_sald, ore_totale,
                source_file, updated_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                r["nr_commessa"],
                r["cliente"],
                r["cliente_norm"],
                r["ore_imba"],
                r["ore_nest"],
                r["ore_pieg"],
                r["ore_prod"],
                r["ore_prog"],
                r["ore_sald"],
                r["ore_totale"],
                source_file,
                now,
            ),
        )
    conn.commit()
    cur.execute("SELECT COUNT(*) AS cnt FROM commesse_ore")
    n = cur.fetchone()['cnt']
    conn.close()
    return {"imported": len(rows), "stored": n, "source_file": source_file}


def match_preventivi_filenames(
    filenames: list[str],
    rows_by_norm: dict[str, list[dict[str, Any]]],
    db_path: str | None = None,
) -> list[dict[str, Any]]:
    """
    Per ogni file offerta:
    1) Se db_path e tabella mapping: nr. preventivo dal nome file → commessa da elenco gestionale → ore.
    2) Altrimenti (fallback): match su cliente_norm.
    Match: mapped | mapped_no_ore | ambiguous_mapping | unique | ambiguous | none | unparsed
    """
    from offerta_commessa_mapping import commesse_ore_by_nr, mapping_by_preventivo

    mp: dict[int, list] = {}
    ore_by_nr: dict[str, dict] = {}
    if db_path:
        try:
            mp = mapping_by_preventivo(db_path)
            ore_by_nr = commesse_ore_by_nr(db_path)
        except sqlite3.OperationalError:
            mp = {}
            ore_by_nr = {}

    results = []
    for fn in filenames:
        p = parse_preventivo_filename(fn)
        if not p:
            results.append(
                {
                    "filename": os.path.basename(fn),
                    "parse_ok": False,
                    "match": "unparsed",
                    "detail": "Nome file non nel formato YYYY_Cliente_preventivo_NNN",
                    "commesse": [],
                }
            )
            continue

        if mp:
            try:
                prev = int(p["numero_preventivo"])
            except (TypeError, ValueError):
                prev = None
            if prev is not None and prev in mp:
                mrows = mp[prev]
                distinct_nc = list(dict.fromkeys([r["nr_commessa"] for r in mrows]))
                rs = mrows[0].get("ragione_sociale") if mrows else None
                if len(distinct_nc) > 1:
                    comm_out = []
                    for nc in distinct_nc:
                        o = ore_by_nr.get(nc)
                        if o:
                            comm_out.append(
                                {
                                    "nr_commessa": nc,
                                    "cliente": o.get("cliente"),
                                    "ore_totale": o.get("ore_totale"),
                                }
                            )
                        else:
                            comm_out.append(
                                {
                                    "nr_commessa": nc,
                                    "cliente": rs,
                                    "ore_totale": None,
                                }
                            )
                    results.append(
                        {
                            "filename": p["filename"],
                            "parse_ok": True,
                            "cliente_da_file": p["cliente_slug"],
                            "cliente_norm": p["cliente_norm"],
                            "numero_preventivo": p["numero_preventivo"],
                            "match": "ambiguous_mapping",
                            "detail": (
                                "Stesso nr. preventivo associato a più commesse nell'elenco importato"
                            ),
                            "commesse": comm_out[:50],
                        }
                    )
                    continue

                nc = distinct_nc[0]
                ore = ore_by_nr.get(nc)
                if ore:
                    results.append(
                        {
                            "filename": p["filename"],
                            "parse_ok": True,
                            "cliente_da_file": p["cliente_slug"],
                            "cliente_norm": p["cliente_norm"],
                            "numero_preventivo": p["numero_preventivo"],
                            "match": "mapped",
                            "detail": "Legame nr. preventivo ↔ commessa (elenco offerte) + ore reparto",
                            "nr_commessa": nc,
                            "ragione_sociale_mapping": rs,
                            "commesse": [
                                {
                                    "nr_commessa": ore["nr_commessa"],
                                    "cliente": ore["cliente"],
                                    "ore_totale": ore["ore_totale"],
                                }
                            ],
                        }
                    )
                    continue

                results.append(
                    {
                        "filename": p["filename"],
                        "parse_ok": True,
                        "cliente_da_file": p["cliente_slug"],
                        "cliente_norm": p["cliente_norm"],
                        "numero_preventivo": p["numero_preventivo"],
                        "match": "mapped_no_ore",
                        "detail": (
                            "Commessa trovata nell'elenco offerte; ore reparto assenti "
                            "in Excel ore per questa commessa"
                        ),
                        "nr_commessa": nc,
                        "ragione_sociale_mapping": rs,
                        "commesse": [
                            {
                                "nr_commessa": nc,
                                "cliente": rs,
                                "ore_totale": None,
                            }
                        ],
                    }
                )
                continue

        norm = p["cliente_norm"]
        commesse = rows_by_norm.get(norm, [])
        if len(commesse) == 0:
            results.append(
                {
                    "filename": p["filename"],
                    "parse_ok": True,
                    "cliente_da_file": p["cliente_slug"],
                    "cliente_norm": norm,
                    "numero_preventivo": p["numero_preventivo"],
                    "match": "none",
                    "detail": "Nessuna commessa con stesso cliente (normalizzato)",
                    "commesse": [],
                }
            )
        elif len(commesse) == 1:
            c = commesse[0]
            results.append(
                {
                    "filename": p["filename"],
                    "parse_ok": True,
                    "cliente_da_file": p["cliente_slug"],
                    "cliente_norm": norm,
                    "numero_preventivo": p["numero_preventivo"],
                    "match": "unique",
                    "detail": "Una sola commessa per questo cliente nel dataset",
                    "commesse": [
                        {
                            "nr_commessa": c["nr_commessa"],
                            "cliente": c["cliente"],
                            "ore_totale": c["ore_totale"],
                        }
                    ],
                }
            )
        else:
            results.append(
                {
                    "filename": p["filename"],
                    "parse_ok": True,
                    "cliente_da_file": p["cliente_slug"],
                    "cliente_norm": norm,
                    "numero_preventivo": p["numero_preventivo"],
                    "match": "ambiguous",
                    "detail": f"Più commesse ({len(commesse)}) con stesso cliente: serve chiave aggiuntiva (es. nr. offerta in anagrafica)",
                    "commesse": [
                        {
                            "nr_commessa": c["nr_commessa"],
                            "cliente": c["cliente"],
                            "ore_totale": c["ore_totale"],
                        }
                        for c in commesse[:50]
                    ],
                }
            )
    return results


def rows_grouped_by_cliente_norm(db_path: str) -> dict[str, list[dict[str, Any]]]:
    conn = _get_db_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT nr_commessa, cliente, cliente_norm, ore_totale,
               ore_imba, ore_nest, ore_pieg, ore_prod, ore_prog, ore_sald
        FROM commesse_ore ORDER BY nr_commessa
        """
    )
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    by_norm: dict[str, list[dict[str, Any]]] = {}
    for r in rows:
        by_norm.setdefault(r["cliente_norm"], []).append(r)
    return by_norm

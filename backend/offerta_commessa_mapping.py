"""
Legame offerta (nr. preventivo) ↔ commessa da Excel gestionale (elenco commesse 2025 rif offerta.xlsm).
"""
from __future__ import annotations

import os
import sqlite3
from datetime import datetime, timezone
from typing import Any

import openpyxl

def default_mapping_xlsx_path() -> str:
    root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    candidates = [
        os.path.join(root, "dati", "elenco commesse 2025 rif offerta.xlsm"),
        os.path.join(
            root,
            "_offerte_extracted",
            "OFFERTE",
            "elenco commesse 2025 rif offerta.xlsm",
        ),
    ]
    for p in candidates:
        if os.path.isfile(p):
            return p
    return candidates[-1]


def _cell_str(v: Any) -> str | None:
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        return v.isoformat()
    s = str(v).strip()
    return s if s else None


def load_rows_from_mapping_xlsx(path: str) -> list[dict[str, Any]]:
    if not os.path.isfile(path):
        raise FileNotFoundError(path)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Sheet1" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Foglio Sheet1 non trovato: {wb.sheetnames}")

    ws = wb["Sheet1"]
    rows_out: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(c is None for c in row):
            continue
        comm = row[9]
        prev_raw = row[11]
        ref = row[10]
        if comm is None or prev_raw is None:
            continue
        try:
            nr_prev = int(float(prev_raw))
        except (TypeError, ValueError):
            continue
        nr_commessa = str(comm).strip()
        rows_out.append(
            {
                "nr_preventivo": nr_prev,
                "nr_commessa": nr_commessa,
                "ragione_sociale": _cell_str(row[7]),
                "riferimento_offerta": _cell_str(ref),
                "data_doc": _cell_str(row[3]),
            }
        )
    wb.close()
    return rows_out


def import_mapping_to_sqlite(
    db_path: str,
    rows: list[dict[str, Any]],
    source_file: str,
) -> dict[str, Any]:
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("DELETE FROM offerta_commessa_map")
    now = datetime.now(timezone.utc).isoformat()
    for r in rows:
        cur.execute(
            """
            INSERT INTO offerta_commessa_map (
                nr_preventivo, nr_commessa, ragione_sociale,
                riferimento_offerta, data_doc, source_file, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                r["nr_preventivo"],
                r["nr_commessa"],
                r["ragione_sociale"],
                r["riferimento_offerta"],
                r["data_doc"],
                source_file,
                now,
            ),
        )
    conn.commit()
    n = cur.execute("SELECT COUNT(*) FROM offerta_commessa_map").fetchone()[0]
    conn.close()
    return {"imported": len(rows), "stored": n, "source_file": source_file}


def mapping_by_preventivo(db_path: str) -> dict[int, list[dict[str, Any]]]:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(
        """
        SELECT nr_preventivo, nr_commessa, ragione_sociale, riferimento_offerta, data_doc
        FROM offerta_commessa_map ORDER BY nr_preventivo, nr_commessa
        """
    )
    by_prev: dict[int, list[dict[str, Any]]] = {}
    for r in cur.fetchall():
        d = dict(r)
        by_prev.setdefault(int(d["nr_preventivo"]), []).append(d)
    conn.close()
    return by_prev


def commesse_ore_by_nr(db_path: str) -> dict[str, dict[str, Any]]:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(
        """
        SELECT nr_commessa, cliente, cliente_norm, ore_totale,
               ore_imba, ore_nest, ore_pieg, ore_prod, ore_prog, ore_sald
        FROM commesse_ore
        """
    )
    out: dict[str, dict[str, Any]] = {}
    for r in cur.fetchall():
        out[r["nr_commessa"]] = dict(r)
    conn.close()
    return out
